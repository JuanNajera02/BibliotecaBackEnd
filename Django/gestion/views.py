from django.http import HttpResponse
from rest_framework import viewsets
from django.db.models import Count
from django.db.models.functions import TruncHour
from django.utils import timezone
from rest_framework.decorators import action
from .models import RDU
from rest_framework.response import Response
from .serializers import RDUSerializer
from .models import Facultad
from .serializers import FacultadSerializer
from .models import Carrera
from .serializers import CarreraSerializer
from .models import Administradores
from .serializers import AdministradoresSerializer
from .models import TipoUsuario
from .serializers import TipoUsuarioSerializer
from .models import Visitias
from .serializers import VisitiasSerializer
from django.db.models import Count, Sum, Case, When, IntegerField
import string
import random
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models.functions import Coalesce
from django.db.models import Q
from django.db.models import OuterRef, Subquery
import openpyxl
from django.shortcuts import get_object_or_404




class RDUViewSet(viewsets.ModelViewSet):
    queryset = RDU.objects.all()
    serializer_class = RDUSerializer

    def create(self, request, *args, **kwargs):
        matricula = request.data.get('matricula', None)

        if matricula is None:
            # Si la matrícula está vacía, genera una matrícula única
            matricula = self.generar_matricula_unico()
            request.data['matricula'] = matricula

        return super().create(request, *args, **kwargs)

    def generar_matricula_unico(self):
        longitud_matricula = 8
        matricula = self.generar_matricula_aleatoria(longitud_matricula)

        # Verificar si la matrícula ya existe en la base de datos
        while RDU.objects.filter(matricula=matricula).exists():
            matricula = self.generar_matricula_aleatoria(longitud_matricula)

        return matricula

    def generar_matricula_aleatoria(self, longitud):
        caracteres = string.ascii_uppercase + string.digits
        return ''.join(random.choice(caracteres) for _ in range(longitud))
    
class FacultadViewSet(viewsets.ModelViewSet):
    queryset = Facultad.objects.all()
    serializer_class = FacultadSerializer

    @action(detail=True, methods=['put'])
    def modificar_nombre(self, request, pk=None):
        facultad = self.get_object()
        nuevo_nombre = request.data.get('nombre', None)

        if nuevo_nombre is None:
            return Response({'error': 'Debes proporcionar un nuevo nombre'}, status=status.HTTP_400_BAD_REQUEST)

        # Modificar el nombre de la facultad y guardar los cambios
        facultad.nombre = nuevo_nombre
        facultad.save()

        # Serializar y devolver la respuesta
        serializer = FacultadSerializer(facultad)
        return Response(serializer.data)



class CarreraViewSet(viewsets.ModelViewSet):
    queryset = Carrera.objects.all()
    serializer_class = CarreraSerializer

    @action(detail=True, methods=['put'])
    def modificar_carrera(self, request, pk=None):
        carrera = self.get_object()

        # Obtener los datos de la solicitud
        nuevo_nombre = request.data.get('nombre', None)
        nueva_facultad_id = request.data.get('facultad', None)

        # Modificar los campos de la carrera si se proporcionan valores
        if nuevo_nombre is not None:
            carrera.nombre = nuevo_nombre
        if nueva_facultad_id is not None:
            carrera.facultad_id = nueva_facultad_id

        # Guardar los cambios
        carrera.save()

        # Serializar y devolver la respuesta
        serializer = CarreraSerializer(carrera)
        return Response(serializer.data)

class AdministradoresViewSet(viewsets.ModelViewSet):
    queryset = Administradores.objects.all()
    serializer_class = AdministradoresSerializer

    @action(detail=False, methods=['POST'])
    def validar_usuario(self, request):
        usuario = request.data.get('usuario', None)
        password = request.data.get('password', None)

        if usuario is None or password is None:
            return Response({'error': 'Debes proporcionar usuario y contraseña'}, status=400)

        try:
            usuario_obj = Administradores.objects.get(usuario=usuario, password=password)
            return Response({'nombre': usuario_obj.nombre, 'mensaje': 'Exito'})
        except Administradores.DoesNotExist:
            return Response({'error': 'Usuario no encontrado o contraseña incorrecta'}, status=404)
    
class TipoUsuarioViewSet(viewsets.ModelViewSet):
    queryset = TipoUsuario.objects.all()
    serializer_class = TipoUsuarioSerializer

    @action(detail=True, methods=['put'])
    def modificar_tipo_usuario(self, request, pk=None):
        tipo_usuario = self.get_object()

        # Obtener el nuevo nombre desde la solicitud
        nuevo_nombre = request.data.get('nombre', None)

        # Modificar el campo 'nombre' si se proporciona un nuevo valor
        if nuevo_nombre is not None:
            tipo_usuario.nombre = nuevo_nombre
            tipo_usuario.save()

            # Serializar y devolver la respuesta
            serializer = TipoUsuarioSerializer(tipo_usuario)
            return Response(serializer.data)

        return Response({'error': 'Debes proporcionar un nuevo nombre'}, status=status.HTTP_400_BAD_REQUEST)

    


class VisitiasViewSet(viewsets.ModelViewSet):
    queryset = Visitias.objects.all()
    serializer_class = VisitiasSerializer

    @action(detail=False, methods=['GET'])
    def obtener_info_por_matricula(self, request):
        matricula = request.query_params.get('matricula', None)

        if matricula is None:
            return Response({'error': 'Debes proporcionar una matrícula'}, status=status.HTTP_400_BAD_REQUEST)

        # Buscar el objeto RDU que tiene la matrícula proporcionada
        rdu_instance = get_object_or_404(RDU, matricula=matricula)

        # Serializar todos los campos de la entidad RDU
        serializer = RDUSerializer(rdu_instance)

        return Response(serializer.data)


    @action(detail=False, methods=['GET'])
    def generarReporteFront(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')

        fecha_inicio = timezone.datetime.strptime(fecha_inicio, '%Y-%m-%d').date() if fecha_inicio else None
        fecha_fin = timezone.datetime.strptime(fecha_fin, '%Y-%m-%d').date() if fecha_fin else None

        if fecha_inicio and fecha_fin:
            visitas_records = Visitias.objects.filter(fechayhora__date__range=[fecha_inicio, fecha_fin])
        else:
            visitas_records = Visitias.objects.all()

        visitas_am = visitas_records.filter(fechayhora__hour__lt=12)
        visitas_pm = visitas_records.filter(fechayhora__hour__gte=12)

        stats_am = self.get_statistics(visitas_am, 'mañana')
        stats_pm = self.get_statistics(visitas_pm, 'tarde')
        stats_general = self.get_statistics(visitas_records, 'general')

        result = self.generate_report(stats_am, stats_pm, stats_general)
        
        workbook = openpyxl.load_workbook('CONCENTRADO_DE_REGISTRO_DIARIO_DE_USUARIOS.xlsx')
        sheet = workbook.active

        # DERECHO					CIENCIA POLITICA					CRIMINALISTICA					DERECHO MIXTO				
        #------------------------ FACULTAD DE DERECHO -----------------------
        #------------------------ DERECHO -----------------------
        CELL_DERECHO_ALUMNOS = sheet.cell(row=8, column=2)
        CELL_DERECHO_DOCENTES = sheet.cell(row=8, column=3)
        CELL_DERECHO_INVESTIGADORES = sheet.cell(row=8, column=4)
        CELL_DERECHO_MUJERES = sheet.cell(row=8, column=5)
        CELL_DERECHO_HOMBRES = sheet.cell(row=8, column=6)

        #------------------------ CIENCIA POLITICA -----------------------
        CELL_CIENCIA_POLITICA_ALUMNOS = sheet.cell(row=8, column=7)
        CELL_CIENCIA_POLITICA_DOCENTES = sheet.cell(row=8, column=8)
        CELL_CIENCIA_POLITICA_INVESTIGADORES = sheet.cell(row=8, column=9)
        CELL_CIENCIA_POLITICA_MUJERES = sheet.cell(row=8, column=10)
        CELL_CIENCIA_POLITICA_HOMBRES = sheet.cell(row=8, column=11)

        #------------------------ CRIMINALISTICA -----------------------
        CELL_CRIMINALISTICA_ALUMNOS = sheet.cell(row=8, column=12)
        CELL_CRIMINALISTICA_DOCENTES = sheet.cell(row=8, column=13)
        CELL_CRIMINALISTICA_INVESTIGADORES = sheet.cell(row=8, column=14)
        CELL_CRIMINALISTICA_MUJERES = sheet.cell(row=8, column=15)
        CELL_CRIMINALISTICA_HOMBRES = sheet.cell(row=8, column=16)

        #------------------------ DERECHO MIXTO -----------------------
        CELL_DERECHO_MIXTO_ALUMNOS = sheet.cell(row=8, column=17)
        CELL_DERECHO_MIXTO_DOCENTES = sheet.cell(row=8, column=18)
        CELL_DERECHO_MIXTO_INVESTIGADORES = sheet.cell(row=8, column=19)
        CELL_DERECHO_MIXTO_MUJERES = sheet.cell(row=8, column=20)
        CELL_DERECHO_MIXTO_HOMBRES = sheet.cell(row=8, column=21)

        #------------------------ MAESTRIA DER -----------------------
        CELL_MAESTRIA_DER_ALUMNOS = sheet.cell(row=8, column=22)
        CELL_MAESTRIA_DER_DOCENTES = sheet.cell(row=8, column=23)
        CELL_MAESTRIA_DER_INVESTIGADORES = sheet.cell(row=8, column=24)
        CELL_MAESTRIA_DER_MUJERES = sheet.cell(row=8, column=25)
        CELL_MAESTRIA_DER_HOMBRES = sheet.cell(row=8, column=26)




        #------------------------ FACULTAD DE INGENIERIA -----------------------
        #------------------------ INGENIERIA  CIVIL-----------------------
        CELL_CIVIL_ALUMNOS = sheet.cell(row=14, column=2)
        CELL_CIVIL_DOCENTES = sheet.cell(row=14, column=3)
        CELL_CIVIL_INVESTIGADORES = sheet.cell(row=14, column=4)
        CELL_CIVIL_MUJERES = sheet.cell(row=14, column=5)
        CELL_CIVIL_HOMBRES = sheet.cell(row=14, column=6)

    
        #------------------------ GEODESIA -----------------------
        CELL_GEODESIA_ALUMNOS = sheet.cell(row=14, column=7)
        CELL_GEODESIA_DOCENTES = sheet.cell(row=14, column=8)
        CELL_GEODESIA_INVESTIGADORES = sheet.cell(row=14, column=9)
        CELL_GEODESIA_MUJERES = sheet.cell(row=14, column=10)
        CELL_GEODESIA_HOMBRES = sheet.cell(row=14, column=11)

        #------------------------ INGENIERIA  SOFTWARE-----------------------
        CELL_SOFTWARE_ALUMNOS = sheet.cell(row=14, column=12)
        CELL_SOFTWARE_DOCENTES = sheet.cell(row=14, column=13)
        CELL_SOFTWARE_INVESTIGADORES = sheet.cell(row=14, column=14)
        CELL_SOFTWARE_MUJERES = sheet.cell(row=14, column=15)
        CELL_SOFTWARE_HOMBRES = sheet.cell(row=14, column=16)

        #------------------------ PROC INDUSTRIALES -----------------------
        CELL_INDUSTRIALES_ALUMNOS = sheet.cell(row=14, column=17)
        CELL_INDUSTRIALES_DOCENTES = sheet.cell(row=14, column=18)
        CELL_INDUSTRIALES_INVESTIGADORES = sheet.cell(row=14, column=19)
        CELL_INDUSTRIALES_MUJERES = sheet.cell(row=14, column=20)
        CELL_INDUSTRIALES_HOMBRES = sheet.cell(row=14, column=21)

        #------------------------ NANO -----------------------
        CELL_NANO_ALUMNOS = sheet.cell(row=14, column=22)
        CELL_NANO_DOCENTES = sheet.cell(row=14, column=23)
        CELL_NANO_INVESTIGADORES = sheet.cell(row=14, column=24)
        CELL_NANO_MUJERES = sheet.cell(row=14, column=25)
        CELL_NANO_HOMBRES = sheet.cell(row=14, column=26)

        #------------------------ MAESTRIA ING ------------------------
        CELL_MAESTRIA_ING_ALUMNOS = sheet.cell(row=14, column=27)
        CELL_MAESTRIA_ING_DOCENTES = sheet.cell(row=14, column=28)
        CELL_MAESTRIA_ING_INVESTIGADORES = sheet.cell(row=14, column=29)
        CELL_MAESTRIA_ING_MUJERES = sheet.cell(row=14, column=30)
        CELL_MAESTRIA_ING_HOMBRES = sheet.cell(row=14, column=31)

        #------------------------ DOCTORADO ING -----------------------
        CELL_DOCTORADO_ING_ALUMNOS = sheet.cell(row=14, column=32)
        CELL_DOCTORADO_ING_DOCENTES = sheet.cell(row=14, column=33)
        CELL_DOCTORADO_ING_INVESTIGADORES = sheet.cell(row=14, column=34)
        CELL_DOCTORADO_ING_MUJERES = sheet.cell(row=14, column=35)
        CELL_DOCTORADO_ING_HOMBRES = sheet.cell(row=14, column=36)

        #------------------------ UNIDAD ACADEMICA DE NEGOCIOS -----------------------
        #------------------------ LRCI -----------------------
        CELL_LRCI_ALUMNOS = sheet.cell(row=20, column=2)
        CELL_LRCI_DOCENTES = sheet.cell(row=20, column=3)
        CELL_LRCI_INVESTIGADORES = sheet.cell(row=20, column=4)
        CELL_LRCI_MUJERES = sheet.cell(row=20, column=5)
        CELL_LRCI_HOMBRES = sheet.cell(row=20, column=6)

        #------------------------ LDEN -----------------------
        CELL_LDEN_ALUMNOS = sheet.cell(row=20, column=7)
        CELL_LDEN_DOCENTES = sheet.cell(row=20, column=8)
        CELL_LDEN_INVESTIGADORES = sheet.cell(row=20, column=9)
        CELL_LDEN_MUJERES = sheet.cell(row=20, column=10)
        CELL_LDEN_HOMBRES = sheet.cell(row=20, column=11)

        #------------------------ MERCADOTECNIA -----------------------
        CELL_MERCADOTECNIA_ALUMNOS = sheet.cell(row=20, column=12)
        CELL_MERCADOTECNIA_DOCENTES = sheet.cell(row=20, column=13)
        CELL_MERCADOTECNIA_INVESTIGADORES = sheet.cell(row=20, column=14)
        CELL_MERCADOTECNIA_MUJERES = sheet.cell(row=20, column=15)
        CELL_MERCADOTECNIA_HOMBRES = sheet.cell(row=20, column=16)

        #------------------------ DISENIO GRAFICO -----------------------
        CELL_DISENIO_ALUMNOS = sheet.cell(row=20, column=17)
        CELL_DISENIO_DOCENTES = sheet.cell(row=20, column=18)
        CELL_DISENIO_INVESTIGADORES = sheet.cell(row=20, column=19)
        CELL_DISENIO_MUJERES = sheet.cell(row=20, column=20)
        CELL_DISENIO_HOMBRES = sheet.cell(row=20, column=21)

        #------------------------ LDEN MIXTO -----------------------
        CELL_LDEN_MIXTO_ALUMNOS = sheet.cell(row=20, column=22)
        CELL_LDEN_MIXTO_DOCENTES = sheet.cell(row=20, column=23)
        CELL_LDEN_MIXTO_INVESTIGADORES = sheet.cell(row=20, column=24)
        CELL_LDEN_MIXTO_MUJERES = sheet.cell(row=20, column=25)
        CELL_LDEN_MIXTO_HOMBRES = sheet.cell(row=20, column=26)

        #------------------------ MAESTRIA NEG -----------------------
        CELL_MAESTRIA_NEG_ALUMNOS = sheet.cell(row=20, column=27)
        CELL_MAESTRIA_NEG_DOCENTES = sheet.cell(row=20, column=28)
        CELL_MAESTRIA_NEG_INVESTIGADORES = sheet.cell(row=20, column=29)
        CELL_MAESTRIA_NEG_MUJERES = sheet.cell(row=20, column=30)
        CELL_MAESTRIA_NEG_HOMBRES = sheet.cell(row=20, column=31)

        #------------------------ DOCTORADO NEG -----------------------
        CELL_DOCTORADO_NEG_ALUMNOS = sheet.cell(row=20, column=32)
        CELL_DOCTORADO_NEG_DOCENTES = sheet.cell(row=20, column=33)
        CELL_DOCTORADO_NEG_INVESTIGADORES = sheet.cell(row=20, column=34)
        CELL_DOCTORADO_NEG_MUJERES = sheet.cell(row=20, column=35)
        CELL_DOCTORADO_NEG_HOMBRES = sheet.cell(row=20, column=36)

        #------------------------ FACULTAD DE ENFERMERIA -----------------------

        #------------------------ ENFERMERIA -----------------------
        CELL_ENFERMERIA_ALUMNOS = sheet.cell(row=26, column=2)
        CELL_ENFERMERIA_DOCENTES = sheet.cell(row=26, column=3)
        CELL_ENFERMERIA_INVESTIGADORES = sheet.cell(row=26, column=4)
        CELL_ENFERMERIA_MUJERES = sheet.cell(row=26, column=5)
        CELL_ENFERMERIA_HOMBRES = sheet.cell(row=26, column=6)


        #------------------------ TECNICO EN ENFERMERIA -----------------------
        CELL_TECNICO_ENFERMERIA_ALUMNOS = sheet.cell(row=26, column=7)
        CELL_TECNICO_ENFERMERIA_DOCENTES = sheet.cell(row=26, column=8)
        CELL_TECNICO_ENFERMERIA_INVESTIGADORES = sheet.cell(row=26, column=9)
        CELL_TECNICO_ENFERMERIA_MUJERES = sheet.cell(row=26, column=10)
        CELL_TECNICO_ENFERMERIA_HOMBRES = sheet.cell(row=26, column=11)

        #------------------------ MAESTRIA ENFERMERIA -----------------------
        CELL_MAESTRIA_ENFERMERIA_ALUMNOS = sheet.cell(row=26, column=12)
        CELL_MAESTRIA_ENFERMERIA_DOCENTES = sheet.cell(row=26, column=13)
        CELL_MAESTRIA_ENFERMERIA_INVESTIGADORES = sheet.cell(row=26, column=14)
        CELL_MAESTRIA_ENFERMERIA_MUJERES = sheet.cell(row=26, column=15)
        CELL_MAESTRIA_ENFERMERIA_HOMBRES = sheet.cell(row=26, column=16)

        #------------------------ DOCTORADO ENFERMERIA -----------------------
        CELL_DOCTORADO_ENFERMERIA_ALUMNOS = sheet.cell(row=26, column=17)
        CELL_DOCTORADO_ENFERMERIA_DOCENTES = sheet.cell(row=26, column=18)
        CELL_DOCTORADO_ENFERMERIA_INVESTIGADORES = sheet.cell(row=26, column=19)
        CELL_DOCTORADO_ENFERMERIA_MUJERES = sheet.cell(row=26, column=20)
        CELL_DOCTORADO_ENFERMERIA_HOMBRES = sheet.cell(row=26, column=21)

        #------------------------ FACULTAD DE CIENCIAS DE LA EDUCACION -----------------------

        #------------------------ CIENCIAS EN EDU -----------------------
        CELL_CIENCIAS_EDU_ALUMNOS = sheet.cell(row=26, column=22)
        CELL_CIENCIAS_EDU_DOCENTES = sheet.cell(row=26, column=23)
        CELL_CIENCIAS_EDU_INVESTIGADORES = sheet.cell(row=26, column=24)
        CELL_CIENCIAS_EDU_MUJERES = sheet.cell(row=26, column=25)
        CELL_CIENCIAS_EDU_HOMBRES = sheet.cell(row=26, column=26)

        #------------------------ EDUCACION MAT -----------------------
        CELL_EDUCACION_MAT_ALUMNOS = sheet.cell(row=26, column=27)
        CELL_EDUCACION_MAT_DOCENTES = sheet.cell(row=26, column=28)
        CELL_EDUCACION_MAT_INVESTIGADORES = sheet.cell(row=26, column=29)
        CELL_EDUCACION_MAT_MUJERES = sheet.cell(row=26, column=30)
        CELL_EDUCACION_MAT_HOMBRES = sheet.cell(row=26, column=31)

        #------------------------ EDUCACION ESP -----------------------
        CELL_EDUCACION_ESP_ALUMNOS = sheet.cell(row=26, column=32)
        CELL_EDUCACION_ESP_DOCENTES = sheet.cell(row=26, column=33)
        CELL_EDUCACION_ESP_INVESTIGADORES = sheet.cell(row=26, column=34)
        CELL_EDUCACION_ESP_MUJERES = sheet.cell(row=26, column=35)
        CELL_EDUCACION_ESP_HOMBRES = sheet.cell(row=26, column=36)

        #------------------------ FACULTAD DE TRABAJO SOCIAL -----------------------
        #------------------------ TRABAJO SOCIAL -----------------------
        CELL_TRABAJO_SOCIAL_ALUMNOS = sheet.cell(row=32, column=2)
        CELL_TRABAJO_SOCIAL_DOCENTES = sheet.cell(row=32, column=3)
        CELL_TRABAJO_SOCIAL_INVESTIGADORES = sheet.cell(row=32, column=4)
        CELL_TRABAJO_SOCIAL_MUJERES = sheet.cell(row=32, column=5)
        CELL_TRABAJO_SOCIAL_HOMBRES = sheet.cell(row=32, column=6)

        #------------------------ TS MIXTO -----------------------
        CELL_TS_MIXTO_ALUMNOS = sheet.cell(row=32, column=7)
        CELL_TS_MIXTO_DOCENTES = sheet.cell(row=32, column=8)
        CELL_TS_MIXTO_INVESTIGADORES = sheet.cell(row=32, column=9)
        CELL_TS_MIXTO_MUJERES = sheet.cell(row=32, column=10)
        CELL_TS_MIXTO_HOMBRES = sheet.cell(row=32, column=11)
        
        #------------------------ FACULTAD DE EDUCACIÓN FISICA Y DEPORTE -----------------------

        #------------------------ EDUCACION FISICA -----------------------
        CELL_EDUCACION_FISICA_ALUMNOS = sheet.cell(row=32, column=17)
        CELL_EDUCACION_FISICA_DOCENTES = sheet.cell(row=32, column=18)
        CELL_EDUCACION_FISICA_INVESTIGADORES = sheet.cell(row=32, column=19)
        CELL_EDUCACION_FISICA_MUJERES = sheet.cell(row=32, column=20)
        CELL_EDUCACION_FISICA_HOMBRES = sheet.cell(row=32, column=21)

        #------------------------ EDU DEPORTE -----------------------
        CELL_EDU_DEPORTE_ALUMNOS = sheet.cell(row=32, column=22)
        CELL_EDU_DEPORTE_DOCENTES = sheet.cell(row=32, column=23)
        CELL_EDU_DEPORTE_INVESTIGADORES = sheet.cell(row=32, column=24)
        CELL_EDU_DEPORTE_MUJERES = sheet.cell(row=32, column=25)
        CELL_EDU_DEPORTE_HOMBRES = sheet.cell(row=32, column=26)

        #------------------------ VARIOS -----------------------

        #------------------------ MEDICINA -----------------------
        CELL_MEDICINA_ALUMNOS = sheet.cell(row=38, column=2)
        CELL_MEDICINA_DOCENTES = sheet.cell(row=38, column=3)
        CELL_MEDICINA_INVESTIGADORES = sheet.cell(row=38, column=4)
        CELL_MEDICINA_MUJERES = sheet.cell(row=38, column=5)
        CELL_MEDICINA_HOMBRES = sheet.cell(row=38, column=6)

        #------------------------ COMUNICACION -----------------------
        CELL_COMUNICACION_ALUMNOS = sheet.cell(row=38, column=7)
        CELL_COMUNICACION_DOCENTES = sheet.cell(row=38, column=8)
        CELL_COMUNICACION_INVESTIGADORES = sheet.cell(row=38, column=9)
        CELL_COMUNICACION_MUJERES = sheet.cell(row=38, column=10)
        CELL_COMUNICACION_HOMBRES = sheet.cell(row=38, column=11)

        #------------------------ INGLES -----------------------
        CELL_INGLES_ALUMNOS = sheet.cell(row=38, column=12)
        CELL_INGLES_DOCENTES = sheet.cell(row=38, column=13)
        CELL_INGLES_INVESTIGADORES = sheet.cell(row=38, column=14)
        CELL_INGLES_MUJERES = sheet.cell(row=38, column=15)
        CELL_INGLES_HOMBRES = sheet.cell(row=38, column=16)

        #------------------------ PREPA CU -----------------------
        CELL_PREPA_CU_ALUMNOS = sheet.cell(row=38, column=17)
        CELL_PREPA_CU_DOCENTES = sheet.cell(row=38, column=18)
        CELL_PREPA_CU_INVESTIGADORES = sheet.cell(row=38, column=19)
        CELL_PREPA_CU_MUJERES = sheet.cell(row=38, column=20)
        CELL_PREPA_CU_HOMBRES = sheet.cell(row=38, column=21)

        #------------------------ OTROS -----------------------
        CELL_OTROS_ALUMNOS = sheet.cell(row=38, column=22)
        CELL_OTROS_DOCENTES = sheet.cell(row=38, column=23)
        CELL_OTROS_INVESTIGADORES = sheet.cell(row=38, column=24)
        CELL_OTROS_MUJERES = sheet.cell(row=38, column=25)
        CELL_OTROS_HOMBRES = sheet.cell(row=38, column=26)

        #------------------------ EXTERNOS -----------------------
        CELL_EXTERNOS_ALUMNOS = sheet.cell(row=38, column=27)
        CELL_EXTERNOS_DOCENTES = sheet.cell(row=38, column=28)
        CELL_EXTERNOS_INVESTIGADORES = sheet.cell(row=38, column=29)
        CELL_EXTERNOS_MUJERES = sheet.cell(row=38, column=30)
        CELL_EXTERNOS_HOMBRES = sheet.cell(row=38, column=31)









        carrera = stats_general['facultad_carrera']
        for i in range(len(carrera)):
            tiposDeUsuario =  carrera[i]['tipos_usuario']
            MUJERES = carrera[i]['mujeres']
            HOMBRES = carrera[i]['hombres']
            ALUMNOS = 0
            DOCENTES = 0
            INVESTIGADORES = 0
            for tipo in tiposDeUsuario:
                print("TIPO " , tipo)
                ALUMNOS = tipo['total'] if tipo['nombre'] == 'Alumno' else ALUMNOS
                INVESTIGADORES = tipo['total'] if tipo['nombre'] == 'Investigador' else INVESTIGADORES
                DOCENTES = tipo['total'] if tipo['nombre'] == 'Docente' else DOCENTES

            if carrera[i]['id_carrera__nombre'] == 'Derecho':
                CELL_DERECHO_ALUMNOS.value = ALUMNOS
                CELL_DERECHO_DOCENTES.value = DOCENTES
                CELL_DERECHO_INVESTIGADORES.value = INVESTIGADORES
                CELL_DERECHO_MUJERES.value = MUJERES
                CELL_DERECHO_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Ciencia politica':
                CELL_CIENCIA_POLITICA_ALUMNOS.value = ALUMNOS
                CELL_CIENCIA_POLITICA_DOCENTES.value = DOCENTES
                CELL_CIENCIA_POLITICA_INVESTIGADORES.value = INVESTIGADORES
                CELL_CIENCIA_POLITICA_MUJERES.value = MUJERES
                CELL_CIENCIA_POLITICA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Criminalistica':
                CELL_CRIMINALISTICA_ALUMNOS.value = ALUMNOS
                CELL_CRIMINALISTICA_DOCENTES.value = DOCENTES
                CELL_CRIMINALISTICA_INVESTIGADORES.value = INVESTIGADORES
                CELL_CRIMINALISTICA_MUJERES.value = MUJERES
                CELL_CRIMINALISTICA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Derecho mixto':
                CELL_DERECHO_MIXTO_ALUMNOS.value = ALUMNOS
                CELL_DERECHO_MIXTO_DOCENTES.value = DOCENTES
                CELL_DERECHO_MIXTO_INVESTIGADORES.value = INVESTIGADORES
                CELL_DERECHO_MIXTO_MUJERES.value = MUJERES
                CELL_DERECHO_MIXTO_HOMBRES.value = HOMBRES   
            elif carrera[i]['id_carrera__nombre'] == 'Maestria der':
                CELL_MAESTRIA_DER_ALUMNOS.value = ALUMNOS
                CELL_MAESTRIA_DER_DOCENTES.value = DOCENTES
                CELL_MAESTRIA_DER_INVESTIGADORES.value = INVESTIGADORES
                CELL_MAESTRIA_DER_MUJERES.value = MUJERES
                CELL_MAESTRIA_DER_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Ingenieria civil':
                CELL_CIVIL_ALUMNOS.value = ALUMNOS
                CELL_CIVIL_DOCENTES.value = DOCENTES
                CELL_CIVIL_INVESTIGADORES.value = INVESTIGADORES
                CELL_CIVIL_MUJERES.value = MUJERES
                CELL_CIVIL_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Geodesia':
                CELL_GEODESIA_ALUMNOS.value = ALUMNOS
                CELL_GEODESIA_DOCENTES.value = DOCENTES
                CELL_GEODESIA_INVESTIGADORES.value = INVESTIGADORES
                CELL_GEODESIA_MUJERES.value = MUJERES
                CELL_GEODESIA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Ingenieria de software':
                CELL_SOFTWARE_ALUMNOS.value = ALUMNOS
                CELL_SOFTWARE_DOCENTES.value = DOCENTES
                CELL_SOFTWARE_INVESTIGADORES.value = INVESTIGADORES
                CELL_SOFTWARE_MUJERES.value = MUJERES
                CELL_SOFTWARE_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Procesos industriales':
                CELL_INDUSTRIALES_ALUMNOS.value = ALUMNOS
                CELL_INDUSTRIALES_DOCENTES.value = DOCENTES
                CELL_INDUSTRIALES_INVESTIGADORES.value = INVESTIGADORES
                CELL_INDUSTRIALES_MUJERES.value = MUJERES
                CELL_INDUSTRIALES_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Nanotecnologia':
                CELL_NANO_ALUMNOS.value = ALUMNOS
                CELL_NANO_DOCENTES.value = DOCENTES
                CELL_NANO_INVESTIGADORES.value = INVESTIGADORES
                CELL_NANO_MUJERES.value = MUJERES
                CELL_NANO_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Maestria ing':
                CELL_MAESTRIA_ING_ALUMNOS.value = ALUMNOS
                CELL_MAESTRIA_ING_DOCENTES.value = DOCENTES
                CELL_MAESTRIA_ING_INVESTIGADORES.value = INVESTIGADORES
                CELL_MAESTRIA_ING_MUJERES.value = MUJERES
                CELL_MAESTRIA_ING_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Doctorado ing':
                CELL_DOCTORADO_ING_ALUMNOS.value = ALUMNOS
                CELL_DOCTORADO_ING_DOCENTES.value = DOCENTES
                CELL_DOCTORADO_ING_INVESTIGADORES.value = INVESTIGADORES
                CELL_DOCTORADO_ING_MUJERES.value = MUJERES
                CELL_DOCTORADO_ING_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Lrci':
                CELL_LRCI_ALUMNOS.value = ALUMNOS
                CELL_LRCI_DOCENTES.value = DOCENTES
                CELL_LRCI_INVESTIGADORES.value = INVESTIGADORES
                CELL_LRCI_MUJERES.value = MUJERES
                CELL_LRCI_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Lden':
                CELL_LDEN_ALUMNOS.value = ALUMNOS
                CELL_LDEN_DOCENTES.value = DOCENTES
                CELL_LDEN_INVESTIGADORES.value = INVESTIGADORES
                CELL_LDEN_MUJERES.value = MUJERES
                CELL_LDEN_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Mercadotecnia':
                CELL_MERCADOTECNIA_ALUMNOS.value = ALUMNOS
                CELL_MERCADOTECNIA_DOCENTES.value = DOCENTES
                CELL_MERCADOTECNIA_INVESTIGADORES.value = INVESTIGADORES
                CELL_MERCADOTECNIA_MUJERES.value = MUJERES
                CELL_MERCADOTECNIA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Diseño grafico':
                CELL_DISENIO_ALUMNOS.value = ALUMNOS
                CELL_DISENIO_DOCENTES.value = DOCENTES
                CELL_DISENIO_INVESTIGADORES.value = INVESTIGADORES
                CELL_DISENIO_MUJERES.value = MUJERES
                CELL_DISENIO_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Lden mixto':
                CELL_LDEN_MIXTO_ALUMNOS.value = ALUMNOS
                CELL_LDEN_MIXTO_DOCENTES.value = DOCENTES
                CELL_LDEN_MIXTO_INVESTIGADORES.value = INVESTIGADORES
                CELL_LDEN_MIXTO_MUJERES.value = MUJERES
                CELL_LDEN_MIXTO_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Maestria neg':
                CELL_MAESTRIA_NEG_ALUMNOS.value = ALUMNOS
                CELL_MAESTRIA_NEG_DOCENTES.value = DOCENTES
                CELL_MAESTRIA_NEG_INVESTIGADORES.value = INVESTIGADORES
                CELL_MAESTRIA_NEG_MUJERES.value = MUJERES
                CELL_MAESTRIA_NEG_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Doctorado neg':
                CELL_DOCTORADO_NEG_ALUMNOS.value = ALUMNOS
                CELL_DOCTORADO_NEG_DOCENTES.value = DOCENTES
                CELL_DOCTORADO_NEG_INVESTIGADORES.value = INVESTIGADORES
                CELL_DOCTORADO_NEG_MUJERES.value = MUJERES
                CELL_DOCTORADO_NEG_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Enfermeria':
                CELL_ENFERMERIA_ALUMNOS.value = ALUMNOS
                CELL_ENFERMERIA_DOCENTES.value = DOCENTES
                CELL_ENFERMERIA_INVESTIGADORES.value = INVESTIGADORES
                CELL_ENFERMERIA_MUJERES.value = MUJERES
                CELL_ENFERMERIA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Tecnico en enfermeria':
                CELL_TECNICO_ENFERMERIA_ALUMNOS.value = ALUMNOS
                CELL_TECNICO_ENFERMERIA_DOCENTES.value = DOCENTES
                CELL_TECNICO_ENFERMERIA_INVESTIGADORES.value = INVESTIGADORES
                CELL_TECNICO_ENFERMERIA_MUJERES.value = MUJERES
                CELL_TECNICO_ENFERMERIA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Maestria enfermeria':
                CELL_MAESTRIA_ENFERMERIA_ALUMNOS.value = ALUMNOS
                CELL_MAESTRIA_ENFERMERIA_DOCENTES.value = DOCENTES
                CELL_MAESTRIA_ENFERMERIA_INVESTIGADORES.value = INVESTIGADORES
                CELL_MAESTRIA_ENFERMERIA_MUJERES.value = MUJERES
                CELL_MAESTRIA_ENFERMERIA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Doctorado enfermeria':
                CELL_DOCTORADO_ENFERMERIA_ALUMNOS.value = ALUMNOS
                CELL_DOCTORADO_ENFERMERIA_DOCENTES.value = DOCENTES
                CELL_DOCTORADO_ENFERMERIA_INVESTIGADORES.value = INVESTIGADORES
                CELL_DOCTORADO_ENFERMERIA_MUJERES.value = MUJERES
                CELL_DOCTORADO_ENFERMERIA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Ciencias en edu':
                CELL_CIENCIAS_EDU_ALUMNOS.value = ALUMNOS
                CELL_CIENCIAS_EDU_DOCENTES.value = DOCENTES
                CELL_CIENCIAS_EDU_INVESTIGADORES.value = INVESTIGADORES
                CELL_CIENCIAS_EDU_MUJERES.value = MUJERES
                CELL_CIENCIAS_EDU_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Educacion mat':
                CELL_EDUCACION_MAT_ALUMNOS.value = ALUMNOS
                CELL_EDUCACION_MAT_DOCENTES.value = DOCENTES
                CELL_EDUCACION_MAT_INVESTIGADORES.value = INVESTIGADORES
                CELL_EDUCACION_MAT_MUJERES.value = MUJERES
                CELL_EDUCACION_MAT_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Educacion esp':
                CELL_EDUCACION_ESP_ALUMNOS.value = ALUMNOS
                CELL_EDUCACION_ESP_DOCENTES.value = DOCENTES
                CELL_EDUCACION_ESP_INVESTIGADORES.value = INVESTIGADORES
                CELL_EDUCACION_ESP_MUJERES.value = MUJERES
                CELL_EDUCACION_ESP_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Trabajo social':
                CELL_TRABAJO_SOCIAL_ALUMNOS.value = ALUMNOS
                CELL_TRABAJO_SOCIAL_DOCENTES.value = DOCENTES
                CELL_TRABAJO_SOCIAL_INVESTIGADORES.value = INVESTIGADORES
                CELL_TRABAJO_SOCIAL_MUJERES.value = MUJERES
                CELL_TRABAJO_SOCIAL_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Ts mixto':
                CELL_TS_MIXTO_ALUMNOS.value = ALUMNOS
                CELL_TS_MIXTO_DOCENTES.value = DOCENTES
                CELL_TS_MIXTO_INVESTIGADORES.value = INVESTIGADORES
                CELL_TS_MIXTO_MUJERES.value = MUJERES
                CELL_TS_MIXTO_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Educacion fisica':
                CELL_EDUCACION_FISICA_ALUMNOS.value = ALUMNOS
                CELL_EDUCACION_FISICA_DOCENTES.value = DOCENTES
                CELL_EDUCACION_FISICA_INVESTIGADORES.value = INVESTIGADORES
                CELL_EDUCACION_FISICA_MUJERES.value = MUJERES
                CELL_EDUCACION_FISICA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Edu deporte':
                CELL_EDU_DEPORTE_ALUMNOS.value = ALUMNOS
                CELL_EDU_DEPORTE_DOCENTES.value = DOCENTES
                CELL_EDU_DEPORTE_INVESTIGADORES.value = INVESTIGADORES
                CELL_EDU_DEPORTE_MUJERES.value = MUJERES
                CELL_EDU_DEPORTE_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Medicina':
                CELL_MEDICINA_ALUMNOS.value = ALUMNOS
                CELL_MEDICINA_DOCENTES.value = DOCENTES
                CELL_MEDICINA_INVESTIGADORES.value = INVESTIGADORES
                CELL_MEDICINA_MUJERES.value = MUJERES
                CELL_MEDICINA_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Comunicacion':
                CELL_COMUNICACION_ALUMNOS.value = ALUMNOS
                CELL_COMUNICACION_DOCENTES.value = DOCENTES
                CELL_COMUNICACION_INVESTIGADORES.value = INVESTIGADORES
                CELL_COMUNICACION_MUJERES.value = MUJERES
                CELL_COMUNICACION_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Ingles':
                CELL_INGLES_ALUMNOS.value = ALUMNOS
                CELL_INGLES_DOCENTES.value = DOCENTES
                CELL_INGLES_INVESTIGADORES.value = INVESTIGADORES
                CELL_INGLES_MUJERES.value = MUJERES
                CELL_INGLES_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Prepa cu':
                CELL_PREPA_CU_ALUMNOS.value = ALUMNOS
                CELL_PREPA_CU_DOCENTES.value = DOCENTES
                CELL_PREPA_CU_INVESTIGADORES.value = INVESTIGADORES
                CELL_PREPA_CU_MUJERES.value = MUJERES
                CELL_PREPA_CU_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Otros':
                CELL_OTROS_ALUMNOS.value = ALUMNOS
                CELL_OTROS_DOCENTES.value = DOCENTES
                CELL_OTROS_INVESTIGADORES.value = INVESTIGADORES
                CELL_OTROS_MUJERES.value = MUJERES
                CELL_OTROS_HOMBRES.value = HOMBRES
            elif carrera[i]['id_carrera__nombre'] == 'Externos':
                CELL_EXTERNOS_ALUMNOS.value = ALUMNOS
                CELL_EXTERNOS_DOCENTES.value = DOCENTES
                CELL_EXTERNOS_INVESTIGADORES.value = INVESTIGADORES
                CELL_EXTERNOS_MUJERES.value = MUJERES
                CELL_EXTERNOS_HOMBRES.value = HOMBRES
            





        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=CONCENTRADO_DE_REGISTRO_DIARIO_DE_USUARIOS.xlsx'
        workbook.save(response)
        
        return response

    def get_statistics(self, queryset, horario):
        tipos_usuario_stats = queryset.values(
            'idRDU__id_carrera__nombre',
            'idRDU__id_carrera__facultad__nombre',
            'idRDU__tipoUsuario__nombre'
        ).annotate(
            total=Count('id')
        ).values(
            'idRDU__id_carrera__nombre',
            'idRDU__id_carrera__facultad__nombre',
            'idRDU__tipoUsuario__nombre',
            'total'
        )

        stats_facultad_carrera = queryset.values(
            'idRDU__id_carrera__nombre',
            'idRDU__id_carrera__facultad__nombre',
        ).annotate(
            total=Count('id'),
            hombres=Coalesce(
                Sum(Case(When(idRDU__sexo='MASCULINO', then=1), default=0, output_field=IntegerField())),
                0
            ),
            mujeres=Coalesce(
                Sum(Case(When(idRDU__sexo='FEMENINO', then=1), default=0, output_field=IntegerField())),
                0
            ),
        )

        return {
            'facultad_carrera': [
                {
                    'id_carrera__nombre': entry['idRDU__id_carrera__nombre'],
                    'id_carrera__facultad__nombre': entry['idRDU__id_carrera__facultad__nombre'],
                    'total': entry['total'],
                    'hombres': entry['hombres'],
                    'mujeres': entry['mujeres'],
                    'tipos_usuario': [
                        {
                            'nombre': tipo_entry['idRDU__tipoUsuario__nombre'],
                            'total': tipo_entry['total'],
                        } for tipo_entry in tipos_usuario_stats
                        if tipo_entry['idRDU__id_carrera__nombre'] == entry['idRDU__id_carrera__nombre'] and
                        tipo_entry['idRDU__id_carrera__facultad__nombre'] == entry['idRDU__id_carrera__facultad__nombre']
                    ],
                } for entry in stats_facultad_carrera
            ],
            'sexo': {
                'hombres': stats_facultad_carrera.aggregate(Sum('hombres'))['hombres__sum'] or 0,
                'mujeres': stats_facultad_carrera.aggregate(Sum('mujeres'))['mujeres__sum'] or 0,
                'total': stats_facultad_carrera.aggregate(Sum('total'))['total__sum'] or 0,
            },
            'horario': horario,
        }

    def generate_report(self, stats_am, stats_pm, stats_general):
        total_tipos_usuario_am = self.get_total_tipos_usuario(stats_am)
        total_tipos_usuario_pm = self.get_total_tipos_usuario(stats_pm)
        total_tipos_usuario_general = self.get_total_tipos_usuario(stats_general)

        # Eliminar la sección "horario" del diccionario
        stats_am.pop('horario', None)
        stats_pm.pop('horario', None)
        stats_general.pop('horario', None)

        # Agregar el total de registros de hombres y mujeres en lugar de "horario"
        total_hombres = (
            stats_am['sexo']['hombres']
            + stats_pm['sexo']['hombres']
            + stats_general['sexo']['hombres']
        )
        total_mujeres = (
            stats_am['sexo']['mujeres']
            + stats_pm['sexo']['mujeres']
            + stats_general['sexo']['mujeres']
        )

        report = {
            'mañana': {
                'stats': stats_am,
                'total_tipos_usuario': total_tipos_usuario_am,
            },
            'tarde': {
                'stats': stats_pm,
                'total_tipos_usuario': total_tipos_usuario_pm,
            },
            'general': {
                'stats': stats_general,
                'total_tipos_usuario': total_tipos_usuario_general,
            }
        }

        return report

    def get_total_tipos_usuario(self, stats):
        total_tipos_usuario = []

        for entry in stats['facultad_carrera']:
            for tipo_usuario in entry['tipos_usuario']:
                tipo_usuario_entry = next(
                    (t for t in total_tipos_usuario if t['nombre'] == tipo_usuario['nombre']),
                    None
                )
                if tipo_usuario_entry:
                    tipo_usuario_entry['total'] += tipo_usuario['total']
                else:
                    total_tipos_usuario.append({
                        'nombre': tipo_usuario['nombre'],
                        'total': tipo_usuario['total'],
                    })

        return total_tipos_usuario
